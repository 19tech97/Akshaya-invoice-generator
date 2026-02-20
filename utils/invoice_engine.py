"""
Invoice generation engine.
Takes configuration + uploaded files and generates one Excel invoice per editor.
Returns a list of generated file paths.
"""

import os
import re
import tempfile
import shutil
import pandas as pd
from num2words import num2words
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from utils.excel_helpers import (
    load_excel_files, get_unique_values, filter_table, select_columns,
    write_to_cell, write_to_cell_copy, write_df_to_excel_location_project,
    set_cell_formats, expand_table_to_last_row, move_and_rename,
)
from utils.formats import (
    level_1_format, level_2_format, level_3_format,
    page_sum_format, amount_sum_format,
    level_1_page_sum_format, level_1_amount_sum_format,
    level_1_description_sum_format,
)


def generate_invoices(config: dict) -> list:
    """
    Main entry point. Accepts a config dict and returns list of generated file paths.

    Required config keys:
        conso_path          - path to the uploaded consolidated Excel
        template_path       - path to the uploaded invoice template
        editor_col          - column name identifying editors
        level_1_header      - column for Type grouping
        level_2_header      - column for Name grouping
        sheet_name          - WS sheet name in template
        inv_sheet_name      - INV sheet name in template
        required_columns    - list of 5 column names for line items
        starting_cell       - e.g. "A3"
        table_name          - name of the Excel Table in the template
        inv_no_cell         - cell for invoice number on INV
        inv_dt_cell         - cell for invoice date on INV
        inv_dt              - invoice date string
        editor_name_cell    - cell for editor name on INV
        inv_total_amt_cell  - cell for amount-in-words on INV
        inv_col_name        - column in conso holding invoice numbers
        pm_info             - dict of PM pricing rules (can be empty)
        currency_prefix     - e.g. "SGD"
        output_dir          - temp directory for output files
    """

    conso_path = config["conso_path"]
    template_path = config["template_path"]
    editor_col = config["editor_col"]
    level_1_header = config["level_1_header"]
    level_2_header = config["level_2_header"]
    ws_sheet = config["sheet_name"]
    inv_sheet = config["inv_sheet_name"]
    req_cols = config["required_columns"]
    starting_cell = config["starting_cell"]
    table_name = config["table_name"]
    inv_no_cell = config["inv_no_cell"]
    inv_dt_cell = config["inv_dt_cell"]
    inv_dt = config["inv_dt"]
    editor_name_cell = config["editor_name_cell"]
    inv_total_amt_cell = config["inv_total_amt_cell"]
    inv_col_name = config["inv_col_name"]
    pm_info = config.get("pm_info", {})
    currency = config.get("currency_prefix", "SGD")
    output_dir = config["output_dir"]

    # Derived cell positions (fixed layout assumptions from original code)
    page_sum_col = "C"
    amount_sum_col = "E"
    inv_des_col = "A"
    inv_page_col = "D"
    inv_amt_col = "E"
    inv_base_row = 12

    # Load data
    conso = load_excel_files(conso_path)
    editors = get_unique_values(table=conso, column_name=editor_col)

    col_letter, base_row = coordinate_from_string(starting_cell)
    col_start = column_index_from_string(col_letter)

    generated = []

    for editor in editors:
        # Create fresh template copy
        invoice = write_to_cell_copy(
            excel_path=template_path, sheet_name=ws_sheet,
            cell=starting_cell, value="", relative=0,
        )

        write_to_cell(invoice, inv_sheet, editor_name_cell, editor)

        # Filter for this editor
        editor_data = filter_table(table=conso, primary_index_col=editor_col, **{editor_col: editor})
        unique_types = get_unique_values(table=editor_data, column_name=level_1_header)

        editor_page_sum = editor_data[req_cols[2]].sum()
        editor_amount_sum = editor_data[req_cols[4]].sum()

        # Amount in words
        amt_words = num2words(editor_amount_sum, to="currency", currency="USD").capitalize()
        amt_words = amt_words.replace("dollars, zero cents", "").replace("dollar, zero cents", "")
        amt_words = amt_words.replace("dollars,", "").replace("dollar,", "")
        # Clean up extra whitespace
        amt_words = " ".join(amt_words.split())

        write_to_cell(invoice, inv_sheet, inv_total_amt_cell, f"{currency} {amt_words} only")
        write_to_cell(invoice, inv_sheet, "D38", editor_page_sum)
        write_to_cell(invoice, inv_sheet, "E38", editor_amount_sum)

        # Invoice number & date
        if inv_col_name and inv_col_name in editor_data.columns:
            inv_no = editor_data[inv_col_name].iloc[0]
            write_to_cell(invoice, inv_sheet, inv_no_cell, inv_no)
        write_to_cell(invoice, inv_sheet, inv_dt_cell, inv_dt)

        current_row = base_row
        inv_current_row = inv_base_row

        # ── Level 1 loop ─────────────────────────────────────────────────
        for level_1 in unique_types:
            write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}", level_1, bold=True)
            set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_1_format)
            current_row += 1

            type_data = filter_table(table=editor_data, **{level_1_header: level_1})
            unique_names = get_unique_values(table=type_data, column_name=level_2_header)
            l1_page_sum = type_data[req_cols[2]].sum()
            l1_amount_sum = type_data[req_cols[4]].sum()

            # INV sheet
            write_to_cell(invoice, inv_sheet, f"{inv_des_col}{inv_current_row}", level_1)
            write_to_cell(invoice, inv_sheet, f"{inv_page_col}{inv_current_row}", l1_page_sum)
            write_to_cell(invoice, inv_sheet, f"{inv_amt_col}{inv_current_row}", l1_amount_sum)
            inv_current_row += 2

            # ── Level 2 loop ─────────────────────────────────────────────
            for level_2 in unique_names:
                final_data = filter_table(table=type_data, **{level_2_header: level_2})
                final_data_copy = final_data.copy()

                po_number_val = final_data["PO number"].iloc[0] if "PO number" in final_data.columns and len(final_data) else None
                po_number = str(po_number_val).strip() if pd.notna(po_number_val) else ""

                po_date_val = final_data["PO approved date"].iloc[0] if "PO approved date" in final_data.columns and len(final_data) else None
                po_date_ts = pd.to_datetime(po_date_val, errors="coerce")
                po_date_str = po_date_ts.strftime("%d-%m-%Y") if not pd.isna(po_date_ts) else ""

                title_first = str(final_data["Title"].iloc[0]) if "Title" in final_data.columns and len(final_data) else level_2

                # ── Branch A: PO + PM pricing ────────────────────────────
                if po_number and level_1 in pm_info:
                    page_value = final_data["Dispatched pages"].iloc[0] if "Dispatched pages" in final_data.columns else 0
                    amount_value = final_data["Amount"].iloc[0] if "Amount" in final_data.columns else 0

                    po_detail = f"(PO no. {po_number}" + (f", dt: {po_date_str}" if po_date_str else "") + ")"

                    write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}", title_first, bold=True)
                    set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_2_format)
                    current_row += 1
                    write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}", po_detail, bold=True)
                    set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_2_format)
                    current_row += 1

                    # Break-up details
                    write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}", "Break up details: ")
                    set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_3_format["A"].copy())
                    write_to_cell(invoice, ws_sheet, f"{page_sum_col}{current_row}", page_value)
                    set_cell_formats(invoice, ws_sheet, f"{page_sum_col}{current_row}", level_3_format["C"].copy())
                    write_to_cell(invoice, ws_sheet, f"{amount_sum_col}{current_row}", amount_value)
                    set_cell_formats(invoice, ws_sheet, f"{amount_sum_col}{current_row}", level_3_format["E"].copy())
                    current_row += 1

                    pmi = pm_info[level_1]
                    write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}",
                                  f"Total page extent = {page_value} Pages")
                    set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_3_format["A"].copy())
                    current_row += 1

                    write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}",
                                  f"Upto {pmi['page_limit']} pages = {currency} {pmi['ad_hoc_till_page_limit']}")
                    set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_3_format["A"].copy())
                    current_row += 1

                    if page_value > pmi["page_limit"] and pmi["after_page_limit_basis"] == "rate":
                        write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}",
                                      f"Additional page @ {currency} {pmi['rate_after_limit']} per page")
                        set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_3_format["A"].copy())
                        current_row += 1
                        excess = page_value - pmi["page_limit"]
                        excess_amt = excess * pmi["rate_after_limit"]
                        write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}",
                                      f"i.e., {excess}*{pmi['rate_after_limit']} = {currency} {excess_amt}")
                        set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_3_format["A"].copy())
                        current_row += 1

                    final_data = final_data.iloc[1:].reset_index(drop=True)

                # ── Branch B: PO, no PM pricing ──────────────────────────
                elif po_number:
                    po_detail = f"(PO no. {po_number}" + (f", dt: {po_date_str}" if po_date_str else "") + ")"
                    final_data = final_data.iloc[1:].reset_index(drop=True)
                    write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}", title_first, bold=True)
                    set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_2_format)
                    current_row += 1
                    write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}", po_detail, bold=True)
                    set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_2_format)
                    current_row += 1

                # ── Branch C: No PO ──────────────────────────────────────
                else:
                    write_to_cell(invoice, ws_sheet, f"{col_letter}{current_row}", level_2, bold=True)
                    set_cell_formats(invoice, ws_sheet, f"{col_letter}{current_row}", level_2_format)
                    current_row += 1

                # Write line-item table
                L3_df = select_columns(table=final_data, required_columns=req_cols)
                page_sum = final_data_copy[req_cols[2]].sum()
                amount_sum_val = final_data_copy[req_cols[4]].sum()

                write_df_to_excel_location_project(
                    df=L3_df, sheet_name=ws_sheet,
                    start_cell=f"{col_letter}{current_row}",
                    excel_path=invoice, relative=0, format_dict=level_3_format,
                )
                current_row += len(L3_df.index)

                # Subtotals
                write_to_cell(invoice, ws_sheet, f"{page_sum_col}{current_row}", page_sum)
                set_cell_formats(invoice, ws_sheet, f"{page_sum_col}{current_row}", page_sum_format)
                write_to_cell(invoice, ws_sheet, f"{amount_sum_col}{current_row}", amount_sum_val)
                set_cell_formats(invoice, ws_sheet, f"{amount_sum_col}{current_row}", amount_sum_format)
                current_row += 2

            # Type total row
            write_to_cell(invoice, ws_sheet, f"A{current_row}", f"{level_1} - Total", bold=True)
            for c in ["A", "B", "D"]:
                set_cell_formats(invoice, ws_sheet, f"{c}{current_row}", level_1_description_sum_format)
            write_to_cell(invoice, ws_sheet, f"{page_sum_col}{current_row}", l1_page_sum, bold=True)
            set_cell_formats(invoice, ws_sheet, f"{page_sum_col}{current_row}", level_1_page_sum_format)
            write_to_cell(invoice, ws_sheet, f"{amount_sum_col}{current_row}", l1_amount_sum, bold=True)
            set_cell_formats(invoice, ws_sheet, f"{amount_sum_col}{current_row}", level_1_amount_sum_format)
            current_row += 2

        # Grand total
        write_to_cell(invoice, ws_sheet, f"A{current_row}", "Grand Total", bold=True)
        for c in ["A", "B", "D"]:
            set_cell_formats(invoice, ws_sheet, f"{c}{current_row}", level_1_description_sum_format)
        write_to_cell(invoice, ws_sheet, f"{page_sum_col}{current_row}", editor_page_sum, bold=True)
        set_cell_formats(invoice, ws_sheet, f"{page_sum_col}{current_row}", level_1_page_sum_format)
        write_to_cell(invoice, ws_sheet, f"{amount_sum_col}{current_row}", editor_amount_sum, bold=True)
        set_cell_formats(invoice, ws_sheet, f"{amount_sum_col}{current_row}", level_1_amount_sum_format)

        # Finalize
        safe_name = re.sub(r'[:*?"<>|\\/]', "-", str(editor)).strip()
        new_filename = f"{safe_name}.xlsx"
        expand_table_to_last_row(invoice, ws_sheet, table_name)
        final_path = move_and_rename(invoice, output_dir, new_filename)
        generated.append(final_path)

    # Post-process: clean up "dollars, zero cents" leftovers
    for fpath in generated:
        try:
            wb = load_workbook(fpath)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            cell.value = cell.value.replace("dollars, zero cents ", "")
                            cell.value = cell.value.replace("dollars, zero cents", "")
                            cell.value = cell.value.replace("dollar, zero cents", "")
            wb.save(fpath)
        except Exception:
            pass

    return generated
