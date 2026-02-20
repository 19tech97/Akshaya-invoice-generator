"""
Excel helper functions — cross-platform version.
Replaces win32com-dependent functions with pure openpyxl equivalents.
PDF conversion is removed (not feasible on a web server without Excel COM).
"""

import os
import shutil
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── Cell formatting helpers ────────────────────────────────────────────────

def set_cell_formats_inplace(ws, cell_ref, format_info: dict):
    """Apply format_info dict to a cell in an already-open worksheet (no save)."""
    cell = ws[cell_ref]

    if "value" in format_info:
        cell.value = format_info["value"]

    if "font" in format_info:
        f, cur = format_info["font"], cell.font
        cell.font = Font(
            name=f.get("name", cur.name),
            sz=f.get("size", cur.sz),
            bold=f.get("bold", cur.bold),
            italic=f.get("italic", cur.italic),
            underline=f.get("underline", cur.underline),
            color=f.get("color", cur.color.rgb if cur.color else None),
        )

    if "fill" in format_info:
        fill = format_info["fill"]
        cell.fill = PatternFill(
            fill_type=fill.get("fill_type"),
            fgColor=fill.get("fgColor"),
            bgColor=fill.get("bgColor"),
        )

    if "alignment" in format_info:
        a = format_info["alignment"]
        cell.alignment = Alignment(
            horizontal=a.get("horizontal"),
            vertical=a.get("vertical"),
            wrap_text=a.get("wrap_text"),
            indent=int(a.get("indent", 0)),
        )

    if "border" in format_info:
        b = format_info["border"]
        cell.border = Border(
            top=Side(style=b.get("top")),
            bottom=Side(style=b.get("bottom")),
            left=Side(style=b.get("left")),
            right=Side(style=b.get("right")),
        )

    if "number_format" in format_info:
        cell.number_format = format_info["number_format"]


def set_cell_formats(excel_path, sheet_name, cell_ref, format_info: dict):
    """Open workbook, apply format to one cell, save."""
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]

    if "value" in format_info:
        cell.value = format_info["value"]

    if "font" in format_info:
        fd = format_info["font"]
        cur = cell.font
        cell.font = Font(
            name=fd.get("name", cur.name),
            sz=fd.get("size", cur.sz),
            bold=fd.get("bold", cur.bold),
            italic=fd.get("italic", cur.italic),
            underline=fd.get("underline", cur.underline),
            color=fd.get("color", cur.color.rgb if cur.color else None),
        )

    if "fill" in format_info:
        fd = format_info["fill"]
        cell.fill = PatternFill(
            fill_type=fd.get("fill_type"),
            fgColor=fd.get("fgColor"),
            bgColor=fd.get("bgColor"),
        )

    if "alignment" in format_info:
        ad = format_info["alignment"]
        cur = cell.alignment
        cell.alignment = Alignment(
            horizontal=ad.get("horizontal", cur.horizontal),
            vertical=ad.get("vertical", cur.vertical),
            wrap_text=ad.get("wrap_text", cur.wrap_text),
            indent=int(ad.get("indent", cur.indent or 0)),
        )

    if "border" in format_info:
        bd = format_info["border"]
        cur = cell.border
        cell.border = Border(
            top=Side(style=bd.get("top", cur.top.style)),
            bottom=Side(style=bd.get("bottom", cur.bottom.style)),
            left=Side(style=bd.get("left", cur.left.style)),
            right=Side(style=bd.get("right", cur.right.style)),
        )

    if "number_format" in format_info:
        cell.number_format = format_info["number_format"]

    wb.save(excel_path)


# ─── Data helpers ────────────────────────────────────────────────────────────

def load_excel_files(file_path: str) -> pd.DataFrame:
    """Load Excel/CSV into DataFrame."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(file_path, header=0)
    elif ext in [".xlsx", ".xls"]:
        engine = "openpyxl" if ext == ".xlsx" else None
        return pd.read_excel(file_path, header=0, engine=engine)
    raise ValueError(f"Unsupported extension: {ext}")


def get_unique_values(table: pd.DataFrame, column_name: str):
    """Return unique non-null values from a column."""
    if column_name not in table.columns:
        raise ValueError(f"Column '{column_name}' not found.")
    return table[column_name].dropna().unique()


def filter_table(table: pd.DataFrame, primary_index_col: str = None, **criteria) -> pd.DataFrame:
    """Filter DataFrame by column==value criteria."""
    df = table.copy()
    for col, val in criteria.items():
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found.")
        df = df[df[col] == val]
    if primary_index_col and primary_index_col in df.columns:
        df = df.set_index([primary_index_col, df.index])
    return df


def select_columns(table: pd.DataFrame, required_columns: list) -> pd.DataFrame:
    """Return DataFrame with only the listed columns."""
    missing = [c for c in required_columns if c not in table.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")
    return table[required_columns].copy()


# ─── Write helpers ───────────────────────────────────────────────────────────

def write_to_cell(excel_path, sheet_name, cell, value, relative=0, bold=False):
    """Write a single value to a cell."""
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    col_letter, row = coordinate_from_string(cell)
    col = column_index_from_string(col_letter)
    row += int(relative)
    target = f"{get_column_letter(col)}{row}"
    cell_obj = ws[target]
    cell_obj.value = value
    if bold:
        cell_obj.font = Font(bold=True)
    wb.save(excel_path)


def write_to_cell_copy(excel_path, sheet_name, cell, value, relative=0):
    """Copy the template, write a value, return the path to the copy."""
    tmp = tempfile.mktemp(suffix=".xlsx")
    shutil.copy(excel_path, tmp)
    write_to_cell(tmp, sheet_name, cell, value, relative)
    return tmp


def write_df_to_excel_location_project(df, excel_path, sheet_name, start_cell,
                                        relative=0, format_dict=None):
    """Write DataFrame rows into Excel with per-column formatting."""
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    col_letter, row_start = coordinate_from_string(start_cell)
    col_start = column_index_from_string(col_letter)
    row_start += int(relative)

    for i, (_, row) in enumerate(df.iterrows()):
        for j, val in enumerate(row):
            cur_row = row_start + i
            cur_col = col_start + j
            cl = get_column_letter(cur_col)
            cell_ref = f"{cl}{cur_row}"
            if format_dict and cl in format_dict:
                fmt = format_dict[cl].copy()
                fmt["value"] = val
                set_cell_formats_inplace(ws, cell_ref, fmt)
            else:
                ws.cell(row=cur_row, column=cur_col, value=val)
    wb.save(excel_path)


def expand_table_to_last_row(file_path, sheet_name, table_name):
    """Resize a named Excel Table so it extends to the last used row."""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    table = ws.tables.get(table_name)
    if table is None:
        # If table not found, silently skip (template might not have one)
        wb.save(file_path)
        return
    start_cell = table.ref.split(":")[0]
    start_col_letter = "".join(filter(str.isalpha, start_cell))
    start_col = ws[start_cell].column
    num_cols = len(table.tableColumns)
    end_col_letter = get_column_letter(start_col + num_cols - 1)
    last_row = ws.max_row
    table.ref = f"{start_col_letter}{int(''.join(filter(str.isdigit, start_cell)))}:{end_col_letter}{last_row}"
    wb.save(file_path)


def move_and_rename(source_path, destination_folder, new_filename):
    """Move file to destination with new name."""
    os.makedirs(destination_folder, exist_ok=True)
    new_path = os.path.join(destination_folder, new_filename)
    if os.path.exists(new_path):
        os.remove(new_path)
    shutil.move(source_path, new_path)
    return new_path
