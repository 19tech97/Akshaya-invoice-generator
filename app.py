"""
Flask web application for invoice generation.
Provides a UI for uploading files, configuring parameters, and downloading invoices.
"""

import os
import uuid
import shutil
import tempfile
import zipfile
import json
import traceback
from flask import Flask, render_template, request, jsonify, send_file, session
from werkzeug.utils import secure_filename

from utils.invoice_engine import generate_invoices
from utils.excel_helpers import load_excel_files

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "invoice-gen-secret-key-change-me")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB max upload
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

# Persistent storage directory (survives across requests)
UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "invoice_gen_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ─── Global error handlers ──────────────────────────────────────────────────
# These ensure ALL errors return JSON (not HTML), preventing the
# "Unexpected token '<'" error in the browser.

@app.errorhandler(404)
def not_found(e):
    if request.path == "/":
        return render_template("index.html")
    return jsonify({"error": "Endpoint not found"}), 404

@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({"error": "Method not allowed"}), 405

@app.errorhandler(413)
def file_too_large(e):
    return jsonify({"error": "File too large. Maximum size is 50 MB."}), 413

@app.errorhandler(500)
def internal_error(e):
    return jsonify({"error": f"Internal server error: {str(e)}"}), 500


def get_session_dir():
    """Get or create a unique temp directory for this session."""
    if "session_id" not in session:
        session["session_id"] = str(uuid.uuid4())
    d = os.path.join(UPLOAD_DIR, session["session_id"])
    os.makedirs(d, exist_ok=True)
    return d


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/health")
def health():
    """Health check endpoint — useful for debugging deployment."""
    return jsonify({"status": "ok", "session_id": session.get("session_id", "none")})


@app.route("/upload-conso", methods=["POST"])
def upload_conso():
    """Upload the consolidated data file and return its column names."""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        f = request.files["file"]
        if not f.filename:
            return jsonify({"error": "Empty filename"}), 400

        sdir = get_session_dir()
        fname = secure_filename(f.filename)
        path = os.path.join(sdir, "conso_" + fname)
        f.save(path)

        df = load_excel_files(path)
        columns = df.columns.tolist()
        session["conso_path"] = path
        session.modified = True
        return jsonify({"columns": columns, "rows": len(df), "filename": fname})
    except Exception as e:
        return jsonify({"error": f"Failed to read file: {str(e)}"}), 400


@app.route("/upload-template", methods=["POST"])
def upload_template():
    """Upload the invoice template file and return its sheet names."""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        f = request.files["file"]
        if not f.filename:
            return jsonify({"error": "Empty filename"}), 400

        sdir = get_session_dir()
        fname = secure_filename(f.filename)
        path = os.path.join(sdir, "template_" + fname)
        f.save(path)

        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        # Get table names (need non-read-only mode)
        wb2 = load_workbook(path)
        tables = {}
        for sn in wb2.sheetnames:
            ws = wb2[sn]
            tables[sn] = list(ws.tables.keys()) if ws.tables else []
        wb2.close()

        session["template_path"] = path
        session.modified = True
        return jsonify({"sheets": sheets, "tables": tables, "filename": fname})
    except Exception as e:
        return jsonify({"error": f"Failed to read template: {str(e)}"}), 400


@app.route("/generate", methods=["POST"])
def generate():
    """Run the invoice generation with the provided configuration."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No configuration data received. Please fill in all steps."}), 400

        # Check that files are still available
        conso_path = session.get("conso_path", "")
        template_path = session.get("template_path", "")

        if not conso_path or not os.path.exists(conso_path):
            return jsonify({"error": "Consolidated data file not found. The server may have restarted. Please go back to Step 1 and re-upload your files."}), 400
        if not template_path or not os.path.exists(template_path):
            return jsonify({"error": "Template file not found. The server may have restarted. Please go back to Step 1 and re-upload your files."}), 400

        sdir = get_session_dir()
        output_dir = os.path.join(sdir, "output")
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir)

        # Parse PM info
        pm_info = {}
        for rule in data.get("pm_rules", []):
            if rule.get("type_name"):
                pm_info[rule["type_name"]] = {
                    "page_limit": float(rule.get("page_limit", 0)),
                    "ad_hoc_till_page_limit": float(rule.get("base_fee", 0)),
                    "after_page_limit_basis": rule.get("overage_basis", "rate"),
                    "rate_after_limit": float(rule.get("overage_rate", 0)),
                }

        config = {
            "conso_path": conso_path,
            "template_path": template_path,
            "editor_col": data["editor_col"],
            "level_1_header": data["level_1_header"],
            "level_2_header": data["level_2_header"],
            "sheet_name": data["sheet_name"],
            "inv_sheet_name": data["inv_sheet_name"],
            "required_columns": data["required_columns"],
            "starting_cell": data.get("starting_cell", "A3"),
            "table_name": data.get("table_name", "invoice"),
            "inv_no_cell": data.get("inv_no_cell", "B3"),
            "inv_dt_cell": data.get("inv_dt_cell", "B5"),
            "inv_dt": data.get("inv_dt", ""),
            "editor_name_cell": data.get("editor_name_cell", "A37"),
            "inv_total_amt_cell": data.get("inv_total_amt_cell", "A40"),
            "inv_col_name": data.get("inv_col_name", ""),
            "pm_info": pm_info,
            "currency_prefix": data.get("currency_prefix", "SGD"),
            "output_dir": output_dir,
        }

        generated = generate_invoices(config)

        files = [os.path.basename(p) for p in generated]
        session["output_dir"] = output_dir
        session.modified = True
        return jsonify({"success": True, "files": files, "count": len(files)})

    except KeyError as e:
        return jsonify({"error": f"Missing required field: {str(e)}. Please complete all steps."}), 400
    except Exception as e:
        tb = traceback.format_exc()
        app.logger.error(f"Generation failed:\n{tb}")
        return jsonify({"error": f"Generation failed: {str(e)}"}), 500


@app.route("/download/<filename>")
def download_file(filename):
    """Download a single generated invoice."""
    try:
        output_dir = session.get("output_dir", "")
        if not output_dir:
            return jsonify({"error": "No files generated yet. Please generate invoices first."}), 404
        path = os.path.join(output_dir, secure_filename(filename))
        if not os.path.exists(path):
            return jsonify({"error": f"File '{filename}' not found. Server may have restarted."}), 404
        return send_file(path, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/download-all")
def download_all():
    """Download all generated invoices as a ZIP."""
    try:
        output_dir = session.get("output_dir", "")
        if not output_dir or not os.path.exists(output_dir):
            return jsonify({"error": "No files generated yet. Please generate invoices first."}), 404

        zip_path = os.path.join(get_session_dir(), "invoices.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname in os.listdir(output_dir):
                fpath = os.path.join(output_dir, fname)
                if os.path.isfile(fpath):
                    zf.write(fpath, fname)

        return send_file(zip_path, as_attachment=True, download_name="invoices.zip")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5000)
